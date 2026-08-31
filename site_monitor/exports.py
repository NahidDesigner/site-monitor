"""Downloadable reports: CSV for anything, xlsx when it should look like a report."""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

BROKEN_COLUMNS = [
    ("domain", "Site"),
    ("page_url", "Page"),
    ("asset_url", "Stylesheet"),
    ("status_code", "Status"),
    ("content_type", "Content type"),
    ("reason", "Why it is broken"),
    ("detected_at", "Detected (UTC)"),
]

PAGESPEED_COLUMNS = [
    ("domain", "Site"),
    ("url", "URL"),
    ("strategy", "Device"),
    ("performance", "Performance"),
    ("lcp_ms", "LCP (ms)"),
    ("cls", "CLS"),
    ("tbt_ms", "TBT (ms)"),
    ("fcp_ms", "FCP (ms)"),
    ("speed_index", "Speed index"),
    ("tti_ms", "TTI (ms)"),
    ("error", "Error"),
    ("tested_at", "Tested (UTC)"),
    ("report_url", "PageSpeed report"),
]

RUNS_COLUMNS = [
    ("id", "Run"),
    ("started_at", "Started (UTC)"),
    ("finished_at", "Finished (UTC)"),
    ("status", "Status"),
    ("sites_checked", "Sites"),
    ("pages_checked", "Pages"),
    ("assets_checked", "Stylesheets"),
    ("broken_assets", "Broken"),
]


def _value(row, key):
    """A cell's value, tolerating a column a given row does not carry.

    Rows come from several queries and columns get added over time; a report
    should not fail to download because one of them predates a column.
    """
    try:
        value = row[key]
    except (KeyError, IndexError):
        return ""
    return "" if value is None else value


def timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def to_csv(rows, columns) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([_value(row, key) for key, _ in columns])
    return buffer.getvalue()


def to_xlsx(rows, columns, *, sheet_title: str) -> bytes:
    """A formatted workbook: frozen header, filters, sane column widths."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="16202B")

    sheet.append([label for _, label in columns])
    for index in range(1, len(columns) + 1):
        cell = sheet.cell(row=1, column=index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    widths = [len(label) + 2 for _, label in columns]
    for row in rows:
        values = []
        for position, (key, _) in enumerate(columns):
            value = _value(row, key)
            values.append(value)
            widths[position] = max(widths[position], min(len(str(value or "")) + 2, 70))
        sheet.append(values)

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{sheet.max_row}"
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
