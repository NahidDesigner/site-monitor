"""Downloadable CSV and Excel reports."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from site_monitor.exports import (
    BROKEN_COLUMNS,
    PAGESPEED_COLUMNS,
    to_csv,
    to_xlsx,
)

BROKEN_ROW = {
    "domain": "dvlfirm.com",
    "page_url": "https://dvlfirm.com/business-law/trust-restatement/",
    "asset_url": "https://dvlfirm.com/wp-content/uploads/elementor/css/post-39321.css?ver=1787903551",
    "status_code": 404,
    "content_type": "text/html",
    "reason": "HTTP 404 (content-type: text/html)",
    "detected_at": "2026-08-29T01:00:00+00:00",
}

PS_ROW = {
    "domain": "a.com", "url": "https://a.com/", "strategy": "mobile",
    "performance": 87.0, "lcp_ms": 2450.7, "cls": 0.021, "tbt_ms": 310.0,
    "fcp_ms": 1200.0, "speed_index": None, "tti_ms": None, "error": None,
    "tested_at": "2026-08-29T01:00:00+00:00",
}


def test_csv_has_friendly_headers_and_the_data():
    output = to_csv([BROKEN_ROW], BROKEN_COLUMNS)
    lines = output.splitlines()

    assert lines[0].startswith("Site,Page,Stylesheet")
    assert "post-39321.css" in lines[1]
    assert "404" in lines[1]


def test_csv_renders_none_as_empty_not_the_word_none():
    output = to_csv([PS_ROW], PAGESPEED_COLUMNS)

    assert "None" not in output
    assert ",," in output  # the empty speed_index / tti columns


def test_csv_quotes_values_containing_commas():
    row = dict(BROKEN_ROW, reason="HTTP 404, served as HTML")

    body = to_csv([row], BROKEN_COLUMNS)

    assert '"HTTP 404, served as HTML"' in body


def test_xlsx_is_a_real_workbook_with_the_data():
    blob = to_xlsx([PS_ROW], PAGESPEED_COLUMNS, sheet_title="PageSpeed")

    sheet = load_workbook(io.BytesIO(blob)).active
    assert sheet.title == "PageSpeed"
    assert sheet.max_row == 2
    assert [cell.value for cell in sheet[1]][:3] == ["Site", "URL", "Device"]
    assert sheet.cell(row=2, column=1).value == "a.com"
    assert sheet.cell(row=2, column=4).value == 87.0


def test_xlsx_freezes_the_header_and_adds_filters():
    """Fifty rows of results are unusable without both."""
    sheet = load_workbook(
        io.BytesIO(to_xlsx([PS_ROW], PAGESPEED_COLUMNS, sheet_title="PageSpeed"))
    ).active

    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


def test_xlsx_handles_an_empty_report():
    sheet = load_workbook(
        io.BytesIO(to_xlsx([], BROKEN_COLUMNS, sheet_title="Broken CSS"))
    ).active

    assert sheet.max_row == 1  # headers only
    assert sheet.auto_filter.ref is None


def test_sheet_titles_are_truncated_to_excels_limit():
    sheet = load_workbook(
        io.BytesIO(to_xlsx([], BROKEN_COLUMNS, sheet_title="x" * 60))
    ).active

    assert len(sheet.title) <= 31
